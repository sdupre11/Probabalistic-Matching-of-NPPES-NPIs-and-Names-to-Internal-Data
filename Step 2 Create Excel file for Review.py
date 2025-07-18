#2 Create Excel file (with conditional formatting) for manual review of unclear matching cases

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# Load matches
matches = pd.read_csv("splink_matched_physicians_multi_state.csv")

# Identify ambiguous matches
many_to_one = matches.groupby('unique_id_l').filter(lambda x: len(x) > 1)
many_to_one['ambiguity_type'] = 'A_to_many_B'
many_to_one['competing_matches_for_A'] = many_to_one.groupby('unique_id_l')['unique_id_r'].transform('count')

one_to_many = matches.groupby('unique_id_r').filter(lambda x: len(x) > 1)
one_to_many['ambiguity_type'] = 'B_to_many_A'
one_to_many['competing_matches_for_B'] = one_to_many.groupby('unique_id_r')['unique_id_l'].transform('count')

ambiguous_matches = pd.concat([many_to_one, one_to_many]).drop_duplicates()

# Create reviewer-friendly layout
reviewer_df = ambiguous_matches.rename(columns={
    'clean_name_l': 'DatasetA_Clean_Name',
    'clean_npi_l': 'DatasetA_Clean_NPI',
    'clean_state_l': 'DatasetA_Clean_State',
    'clean_name_r': 'DatasetB_Clean_Name',
    'clean_npi_r': 'DatasetB_Clean_NPI',
    'all_states_r': 'DatasetB_All_States'
})

reviewer_df['Decision'] = ""  # Add empty column for reviewer decisions

# Save to Excel
excel_file = "splink_ambiguous_matches_reviewer_friendly.xlsx"
reviewer_df.to_excel(excel_file, index=False)

# Apply conditional formatting
wb = load_workbook(excel_file)
ws = wb.active

yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

# Highlight competing matches
col_compete_A = ws['J'][0].column_letter
col_compete_B = ws['K'][0].column_letter

ws.conditional_formatting.add(
    f"{col_compete_A}2:{col_compete_A}{ws.max_row}",
    CellIsRule(operator="greaterThan", formula=["2"], fill=red_fill)
)
ws.conditional_formatting.add(
    f"{col_compete_B}2:{col_compete_B}{ws.max_row}",
    CellIsRule(operator="greaterThan", formula=["2"], fill=red_fill)
)

wb.save(excel_file)
print(f"📊 Reviewer-friendly Excel saved: '{excel_file}'")